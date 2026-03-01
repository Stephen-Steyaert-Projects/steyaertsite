from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import render, redirect, get_object_or_404
from .forms import AddMovieForm, RandomMovieForm
from random import sample, shuffle
from .models import Movie, RandomMovieResult
from django.contrib import messages
from django.utils import timezone
from datetime import timedelta
import csv
import io
from openpyxl import load_workbook


def index(request):
    if request.user.is_authenticated:
        return redirect("home")
    return render(request, "moviedb/index.html")


@login_required(login_url="/auth/login")
def home(request):
    return render(request, "moviedb/home.html")


@login_required(login_url="/auth/login")
def all_movies(request):
    g_movies = (
        Movie.objects.filter(rating="G")
        .order_by("title")
        .values("title")
        .distinct()
    )

    pg_movies = (
        Movie.objects.filter(rating="PG")
        .order_by("title")
        .values("title")
        .distinct()
    )

    pg13_movies = (
        Movie.objects.filter(rating="PG-13")
        .order_by("title")
        .values("title")
        .distinct()
    )

    r_movies = (
        Movie.objects.filter(rating="R")
        .order_by("title")
        .values("title")
        .distinct()
    )

    nr_movies = (
        Movie.objects.filter(rating="NR")
        .order_by("title")
        .values("title")
        .distinct()
    )

    tv_movies = (
        Movie.objects.filter(rating="TV")
        .order_by("title")
        .values("title")
        .distinct()
    )

    ctx = {
        "g_movies": g_movies if g_movies else [],
        "pg_movies": pg_movies if pg_movies else [],
        "pg13_movies": pg13_movies if pg13_movies else [],
        "r_movies": r_movies if r_movies else [],
        "nr_movies": nr_movies if nr_movies else [],
        "tv_movies": tv_movies if tv_movies else [],
    }

    return render(request, "moviedb/all_movies.html", ctx)


@login_required(login_url="/auth/login")
def movies_by_rating(request, rating):
    # Map URL rating slug to database rating and template details
    rating_config = {
        "g": {"db_rating": "G", "var_name": "g_movies", "template": "g_movies.html"},
        "pg": {"db_rating": "PG", "var_name": "pg_movies", "template": "pg_movies.html"},
        "pg-13": {"db_rating": "PG-13", "var_name": "pg13_movies", "template": "pg13_movies.html"},
        "r": {"db_rating": "R", "var_name": "r_movies", "template": "r_movies.html"},
        "nr": {"db_rating": "NR", "var_name": "nr_movies", "template": "nr_movies.html"},
        "tv": {"db_rating": "TV", "var_name": "tv_movies", "template": "tv_shows.html"},
    }

    config = rating_config.get(rating.lower())
    if not config:
        return redirect("all")

    movies = (
        Movie.objects.filter(rating=config["db_rating"])
        .order_by("title")
        .values("title")
        .distinct()
    )

    ctx = {config["var_name"]: movies}
    return render(request, f"moviedb/{config['template']}", ctx)


@login_required(login_url="/auth/login")
def add(request):
    if request.method == "POST":
        form = AddMovieForm(request.POST)
        if form.is_valid():
            # Check for duplicate before saving
            title = form.cleaned_data["title"]
            rating = form.cleaned_data["rating"]
            disk = form.cleaned_data["disk"]

            if Movie.objects.filter(title=title, rating=rating, disk=disk).exists():
                # Get the display value for the disk type (e.g., "DVD" instead of "dvd")
                disk_display = Movie(disk=disk).get_disk_display()
                messages.warning(
                    request,
                    f"The movie '{title}' ({rating}, {disk_display}) is already in the Steyaert Movie Database.",
                )
                return redirect("add")
            else:
                form.save()
                messages.success(request, f"'{title}' has been added to the Steyaert Movie Database!")
                return redirect("add")
    else:
        form = AddMovieForm()

    return render(request, "moviedb/add.html", {"form": form})


@login_required(login_url="/auth/login")
def search_results(request, title: str):
    _title = title.strip()
    if not _title or _title.lower() == "title":
        ctx = {"titles": ["Invalid Search"], "t": _title}
    else:
        results = Movie.objects.filter(title__icontains=_title)
        if results:
            ctx = {"titles": results, "t": _title}
        else:
            ctx = {
                "titles": [],
                "t": _title,
            }
    return render(request, "moviedb/search_results.html", ctx)


@login_required(login_url="/auth/login")
def random(request):
    form = RandomMovieForm()
    return render(request, "moviedb/random.html", {"form": form})


@login_required(login_url="/auth/login")
def random_results(request, result_id=None):
    # Purge old results (>7 days) on-the-fly
    seven_days_ago = timezone.now() - timedelta(days=7)
    RandomMovieResult.objects.filter(created_at__lt=seven_days_ago).delete()

    if request.method == "POST":
        form = RandomMovieForm(request.POST)
        if form.is_valid():
            num_movies = form.cleaned_data["movies"]
            selected_ratings = form.cleaned_data["ratings"]

            movies = (
                Movie.objects.filter(rating__in=selected_ratings)
                .order_by("title")
                .values("title", "rating")
                .distinct()
            )

            movie_list = [{"title": m["title"], "rating": m["rating"]} for m in movies]
            shuffle(movie_list)

            if len(movie_list) < num_movies:
                num_movies = len(movie_list)

            generated = sample(movie_list, num_movies)

            # Store in database instead of session
            result = RandomMovieResult.objects.create(movies=generated)

            # Redirect to shareable URL
            return redirect("random_results", result_id=result.id)

    # GET request - display results by UUID
    if result_id:
        result = get_object_or_404(RandomMovieResult, id=result_id)
        return render(request, "moviedb/random_results.html", {"generated": result.movies})

    # No result_id → back to generator
    return redirect("random_movie_generator")




def is_admin(user):
    return user.is_staff


@login_required(login_url="/auth/login")
@user_passes_test(is_admin)
def add_mass_movies(request):
    if request.method == "POST" and request.FILES.get("csv_file"):
        uploaded_file = request.FILES["csv_file"]
        file_name = uploaded_file.name.lower()

        if not (file_name.endswith(".csv") or file_name.endswith(".xlsx")):
            messages.error(request, "Please upload a valid .csv or .xlsx file.")
            return redirect("add_mass")

        try:
            valid_ratings = {
                choice[0].upper() for choice in Movie._meta.get_field("rating").choices
            }
            valid_disks = {
                choice[0].lower() for choice in Movie._meta.get_field("disk").choices
            }

            rows = []
            if file_name.endswith(".csv"):
                raw_data = uploaded_file.read()
                # Try UTF-8 with BOM first (Excel on Windows), then UTF-8, then fallback to latin-1
                try:
                    decoded_file = raw_data.decode("utf-8-sig")
                except UnicodeDecodeError:
                    try:
                        decoded_file = raw_data.decode("utf-8")
                    except UnicodeDecodeError:
                        decoded_file = raw_data.decode("latin-1")
                io_string = io.StringIO(decoded_file)
                reader = csv.reader(io_string)
                next(reader)  # Skip header row
                rows = list(reader)
            elif file_name.endswith(".xlsx"):
                workbook = load_workbook(uploaded_file)
                # Always use the first sheet
                sheet = workbook.worksheets[0]
                rows = list(sheet.iter_rows(min_row=2, values_only=True))

            count = 0
            for row in rows:
                if not row:
                    continue

                # Handle both formats: 3-column (title,rating,disk) or 5-column (title,rated,4k,blu-ray,dvd)
                if len(row) == 5:
                    # Excel format: Title, Rated, 4k, Blu-Ray, DVD
                    title = str(row[0]).strip() if row[0] else ""
                    rating = str(row[1]).strip() if row[1] else ""

                    # Check which disk format has an X
                    disk = None
                    if row[2] and str(row[2]).strip().upper() == "X":
                        disk = "4k"
                    elif row[3] and str(row[3]).strip().upper() == "X":
                        disk = "blu-ray"
                    elif row[4] and str(row[4]).strip().upper() == "X":
                        disk = "dvd"

                    if not disk:
                        messages.warning(request, f"Skipping row with no disk format marked: {title}")
                        continue

                elif len(row) == 3:
                    # CSV format: title, rating, disk
                    title, rating, disk = [str(item).strip() if item else "" for item in row]
                    disk = disk.lower()
                else:
                    messages.warning(request, f"Skipping invalid row (expected 3 or 5 columns): {row}")
                    continue

                if not title or not rating:
                    messages.warning(request, f"Skipping row with empty values: {row}")
                    continue

                rating = rating.upper()

                if Movie.objects.filter(title=title, rating=rating, disk=disk).exists():
                    disk_display = Movie(disk=disk).get_disk_display()
                    messages.info(
                        request, f"Skipping duplicate movie: {title} ({rating}, {disk_display})"
                    )
                    continue

                if rating not in valid_ratings:
                    messages.warning(
                        request, f"Invalid rating '{rating}' in row: {row}"
                    )
                    continue

                if disk not in valid_disks:
                    messages.warning(
                        request, f"Invalid disk type '{disk}' in row: {row}"
                    )
                    continue

                Movie.objects.create(title=title, rating=rating, disk=disk)
                count += 1

            messages.success(request, f"Successfully added {count} movies.")
        except Exception as e:
            messages.error(request, f"Error processing file: {e}")

        return redirect("add_mass")

    return render(request, "moviedb/mass_add.html")
