from django.shortcuts import render, redirect, get_object_or_404
from django.core.cache import cache
from django.http import JsonResponse, HttpResponse
from io import BytesIO
from PIL import Image
from django.core.files.base import ContentFile
import openpyxl

def _clear_set_caches():
    from django.contrib.auth.models import User
    user_ids = User.objects.values_list('id', flat=True)
    cache.delete_many([f'home_{uid}' for uid in user_ids])


def _clear_user_caches():
    from django.contrib.auth.models import User
    user_ids: list[int] = list(User.objects.values_list('id', flat=True))
    keys: list[str] = []
    for uid in user_ids:
        keys += [f'home_{uid}', f'owned_{uid}', f'missing_{uid}', f'edit_collection_{uid}']
    cache.delete_many(keys)


def _convert_set_image(instance, old_image=None):
    original_path = instance.image.path
    img = Image.open(original_path).convert('RGB')
    output = BytesIO()
    img.save(output, format='WEBP', quality=85)
    output.seek(0)
    filename = instance.name.lower().replace(' ', '_') + '.webp'
    instance.image.delete(save=False)
    if old_image:
        old_image.delete(save=False)
    instance.image.save(filename, ContentFile(output.read()), save=True)

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from .forms import CardForm, SetForm
from django.core.exceptions import PermissionDenied, ValidationError
from django.db.models import Case, When, Value, IntegerField, Prefetch, Count, Q, Sum
from .models import Set, Card, OwnedCard, Deck, DeckCard

# Create your views here.
def index(request):
    if request.user.is_authenticated:
        return redirect('home')
    return render(request, 'swccgdb/index.html')

def all_cards(request):
    sets = cache.get('all_cards_data')
    if sets is None:
        card_order = Card.objects.annotate(
            side_order=_side_order(),
            type_order=_card_type_order(),
        ).order_by('side_order', 'type_order', 'name')
        sets = list(Set.objects.prefetch_related(
            Prefetch('cards', queryset=card_order)
        ).order_by('name'))
        cache.set('all_cards_data', sets, 86400)
    return render(request, 'swccgdb/all-cards.html', {'sets': sets})


@login_required
def home(request):
    cache_key = f'home_{request.user.id}'
    data = cache.get(cache_key)
    if data is None:
        sets = Set.objects.annotate(
            total=Count('cards', distinct=True),
            owned=Count('cards', distinct=True, filter=Q(
                cards__owned_by__user=request.user,
            ) & (Q(cards__owned_by__copies_bb__gt=0) | Q(cards__owned_by__copies_wb__gt=0))),
        ).order_by('released', 'name')
        for s in sets:
            s.pct = round(s.owned / s.total * 100) if s.total else 0
        data = {
            'sets': list(sets),
            'total_cards': sum(s.total for s in sets),
            'total_owned': sum(s.owned for s in sets),
        }
        cache.set(cache_key, data, 86400)
    return render(request, 'swccgdb/home.html', data)

def _card_type_order(prefix=''):
    f = f'{prefix}card_type' if prefix else 'card_type'
    return Case(
        When(**{f: 'admirals_order'}, then=Value(0)),
        When(**{f: 'alien_character'}, then=Value(1)),
        When(**{f: 'rebel_character'}, then=Value(2)),
        When(**{f: 'droid_character'}, then=Value(3)),
        When(**{f: 'imperial_character'}, then=Value(4)),
        When(**{f: 'device'}, then=Value(5)),
        When(**{f: 'effect'}, then=Value(6)),
        When(**{f: 'epic_event'}, then=Value(7)),
        When(**{f: 'interrupt'}, then=Value(8)),
        When(**{f: 'jedi_test'}, then=Value(9)),
        When(**{f: 'location'}, then=Value(10)),
        When(**{f: 'objective'}, then=Value(11)),
        When(**{f: 'starship'}, then=Value(12)),
        When(**{f: 'vehicle'}, then=Value(13)),
        When(**{f: 'weapon'}, then=Value(14)),
        default=Value(15),
        output_field=IntegerField(),
    )

def _side_order(prefix=''):
    f = f'{prefix}side' if prefix else 'side'
    return Case(
        When(**{f: 'L'}, then=Value(0)),
        When(**{f: 'D'}, then=Value(1)),
        default=Value(2),
        output_field=IntegerField(),
    )

@login_required
def add_copy(request, card_id: int, border: str):
    if request.method != 'POST' or border not in ('bb', 'wb'):
        return JsonResponse({'error': 'Invalid request'}, status=400)
    card = get_object_or_404(Card, id=card_id)
    owned, _ = OwnedCard.objects.get_or_create(user=request.user, card=card)
    if border == 'bb':
        owned.copies_bb += 1
    else:
        owned.copies_wb += 1
    owned.save()
    return JsonResponse({'copies_bb': owned.copies_bb, 'copies_wb': owned.copies_wb, 'copies': owned.copies})


@login_required
def export_collection_by_set(request):
    owned_map = {
        oc.card_id: oc
        for oc in OwnedCard.objects.filter(user=request.user)
    }
    sets = Set.objects.prefetch_related('cards').order_by('released', 'name')

    wb = openpyxl.Workbook()
    if wb.active:
        wb.remove(wb.active)

    for card_set in sets:
        ws = wb.create_sheet(title=card_set.name[:31])
        ws.append(['Card ID', 'Name', 'Type', 'Side', 'Rarity', 'BB', 'WB'])
        ws.column_dimensions['A'].hidden = True
        for card in card_set.cards.order_by('name'):
            oc = owned_map.get(card.id)
            ws.append([
                card.id,
                card.name,
                card.get_card_type_display(),
                card.get_side_display(),
                card.rarity,
                oc.copies_bb if oc else 0,
                oc.copies_wb if oc else 0,
            ])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="swccg-collection-by-set.xlsx"'
    return response


@login_required
def import_collection_by_set(request):
    if request.method != 'POST' or 'file' not in request.FILES:
        return redirect('owned_cards')

    wb = openpyxl.load_workbook(request.FILES['file'])
    updated = 0

    for ws in wb.worksheets:
        set_name = ws.title
        try:
            card_set = Set.objects.get(name=set_name)
        except Set.DoesNotExist:
            continue

        for row in ws.iter_rows(min_row=2, values_only=True):
            card_id, card_name, _, _, _, copies_bb, copies_wb = row
            if not card_id and not card_name:
                continue
            try:
                copies_bb = int(copies_bb or 0)
                copies_wb = int(copies_wb or 0)
                if card_id:
                    card = Card.objects.get(id=card_id, card_set=card_set)
                else:
                    card = Card.objects.get(name=card_name, card_set=card_set)
                owned, _ = OwnedCard.objects.get_or_create(user=request.user, card=card)
                owned.copies_bb = copies_bb
                owned.copies_wb = copies_wb
                owned.save()
                updated += 1
            except (Card.DoesNotExist, ValueError, TypeError):
                continue

    messages.success(request, f'Imported {updated} cards successfully.')
    return redirect('owned_cards')


@login_required
def remove_copy(request, card_id: int, border: str):
    if request.method != 'POST' or border not in ('bb', 'wb'):
        return JsonResponse({'error': 'Invalid request'}, status=400)
    card = get_object_or_404(Card, id=card_id)
    try:
        owned = OwnedCard.objects.get(user=request.user, card=card)
        if border == 'bb' and owned.copies_bb > 0:
            owned.copies_bb -= 1
        elif border == 'wb' and owned.copies_wb > 0:
            owned.copies_wb -= 1
        owned.save()
        return JsonResponse({'copies_bb': owned.copies_bb, 'copies_wb': owned.copies_wb, 'copies': owned.copies})
    except OwnedCard.DoesNotExist:
        return JsonResponse({'copies_bb': 0, 'copies_wb': 0, 'copies': 0})


@login_required
def save_collection(request):
    if request.method != 'POST':
        return redirect('owned_cards')
    for key, value in request.POST.items():
        if key.startswith('bb_'):
            card_id = int(key[3:])
            wb_key = f'wb_{card_id}'
            try:
                bb = max(0, int(value or 0))
                wb = max(0, int(request.POST.get(wb_key, 0) or 0))
                card = Card.objects.get(id=card_id)
                owned, _ = OwnedCard.objects.get_or_create(user=request.user, card=card)
                owned.copies_bb = bb
                owned.copies_wb = wb
                owned.save()
            except (Card.DoesNotExist, ValueError, TypeError):
                continue
    cache.delete(f'home_{request.user.id}')
    cache.delete(f'owned_{request.user.id}')
    cache.delete(f'missing_{request.user.id}')
    cache.delete(f'edit_collection_{request.user.id}')
    next_url = request.POST.get('next', 'owned_cards')
    return redirect(next_url)


@staff_member_required
def edit_card(request, card_id: int):
    card = get_object_or_404(Card, id=card_id)
    if request.method == 'POST':
        form = CardForm(request.POST, instance=card)
        if form.is_valid():
            form.save()
            cache.delete('all_cards_data')
            messages.success(request, f'"{card.name}" updated successfully.')
            return redirect('edit_card', card_id=card.id)
    else:
        form = CardForm(instance=card)
    return render(request, 'swccgdb/edit-card.html', {'form': form, 'card': card})


@staff_member_required
def add_card(request):
    if request.method == 'POST':
        form = CardForm(request.POST)
        if form.is_valid():
            card = form.save()
            cache.delete('all_cards_data')
            _clear_user_caches()
            messages.success(request, f'"{card.name}" added successfully.')
            return redirect('add_card')
    else:
        form = CardForm()
    return render(request, 'swccgdb/add-card.html', {'form': form})


@staff_member_required
def sets_list(request):
    sets = Set.objects.order_by('released', 'name')
    return render(request, 'swccgdb/sets-list.html', {'sets': sets})


@staff_member_required
def add_set(request):
    if request.method == 'POST':
        form = SetForm(request.POST, request.FILES)
        if form.is_valid():
            instance = form.save()
            if instance.image:
                _convert_set_image(instance)
            _clear_set_caches()
            messages.success(request, f'"{instance.name}" added successfully.')
            return redirect('add_set')
    else:
        form = SetForm()
    return render(request, 'swccgdb/add-set.html', {'form': form})


@staff_member_required
def edit_set(request, set_id: int):
    card_set = get_object_or_404(Set, id=set_id)
    if request.method == 'POST':
        old_image = card_set.image if card_set.image else None
        form = SetForm(request.POST, request.FILES, instance=card_set)
        if form.is_valid():
            if form.cleaned_data.get('image') is False and old_image:
                old_image.delete(save=False)
            instance = form.save()
            if 'image' in request.FILES and instance.image:
                _convert_set_image(instance, old_image=old_image)
            _clear_set_caches()
            messages.success(request, f'"{instance.name}" updated successfully.')
            return redirect('edit_set', set_id=instance.id)
    else:
        form = SetForm(instance=card_set)
    return render(request, 'swccgdb/edit-set.html', {'form': form, 'card_set': card_set})


@login_required
def edit_collection(request):
    cache_key = f'edit_collection_{request.user.id}'
    data = cache.get(cache_key)
    if data is None:
        owned_map = {
            oc.card_id: oc
            for oc in OwnedCard.objects.filter(user=request.user)
        }
        cards = (
            Card.objects
            .select_related('card_set')
            .annotate(side_order=_side_order(), type_order=_card_type_order())
            .order_by('card_set__name', 'side_order', 'type_order', 'name')
        )
        rows = []
        for card in cards:
            oc = owned_map.get(card.id)
            rows.append({
                'card': card,
                'bb': oc.copies_bb if oc else 0,
                'wb': oc.copies_wb if oc else 0,
            })
        all_sets = list(_all_sets())
        data = {'rows': rows, 'all_sets': all_sets}
        cache.set(cache_key, data, 86400)
    return render(request, 'swccgdb/edit-collection.html', data)


def _all_sets():
    return Set.objects.order_by('released', 'name').values_list('name', flat=True)


@login_required
def owned_cards(request):
    cache_key = f'owned_{request.user.id}'
    data = cache.get(cache_key)
    if data is None:
        owned = list(
            OwnedCard.objects
            .filter(user=request.user)
            .filter(Q(copies_bb__gt=0) | Q(copies_wb__gt=0))
            .select_related('card', 'card__card_set')
            .annotate(side_order=_side_order('card__'), type_order=_card_type_order('card__'))
            .order_by('card__card_set__name', 'side_order', 'type_order', 'card__name')
        )
        data = {'owned': owned, 'all_sets': list(_all_sets())}
        cache.set(cache_key, data, 86400)
    return render(request, 'swccgdb/owned.html', data)


@login_required
def missing_cards(request):
    cache_key = f'missing_{request.user.id}'
    data = cache.get(cache_key)
    if data is None:
        owned_ids = OwnedCard.objects.filter(user=request.user).filter(
            Q(copies_bb__gt=0) | Q(copies_wb__gt=0)
        ).values_list('card_id', flat=True)
        missing = list(
            Card.objects
            .exclude(id__in=owned_ids)
            .select_related('card_set')
            .annotate(side_order=_side_order(), type_order=_card_type_order())
            .order_by('card_set__name', 'side_order', 'type_order', 'name')
        )
        data = {'missing': missing, 'all_sets': list(_all_sets())}
        cache.set(cache_key, data, 86400)
    return render(request, 'swccgdb/missing.html', data)


# --- Deck views ---

@login_required
def deck_list(request):
    my_decks = Deck.objects.filter(user=request.user).annotate(
        card_count=Sum('deck_cards__quantity')
    ).order_by('name')
    saved_decks = Deck.objects.filter(saved_by=request.user).annotate(
        card_count=Sum('deck_cards__quantity')
    ).order_by('name')
    return render(request, 'swccgdb/decks/list.html', {
        'my_decks': my_decks,
        'saved_decks': saved_decks,
    })


@login_required
def deck_new(request):
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if name:
            deck = Deck.objects.create(user=request.user, name=name)
            return redirect('deck_edit', deck_id=deck.id)
        messages.error(request, 'A name is required.')
    return render(request, 'swccgdb/decks/new.html')


def deck_detail(request, deck_id):
    deck = get_object_or_404(Deck, id=deck_id)
    is_owner = request.user.is_authenticated and deck.user == request.user
    if not is_owner and not deck.is_public:
        raise PermissionDenied
    deck_cards = (
        deck.deck_cards
        .select_related('owned_card__card__card_set')
        .annotate(
            type_order=_card_type_order('owned_card__card__'),
            side_order=_side_order('owned_card__card__'),
        )
        .order_by('side_order', 'type_order', 'owned_card__card__name')
    )
    total = deck_cards.aggregate(t=Sum('quantity'))['t'] or 0
    is_saved = (
        request.user.is_authenticated
        and not is_owner
        and deck.saved_by.filter(id=request.user.id).exists()
    )
    return render(request, 'swccgdb/decks/detail.html', {
        'deck': deck,
        'deck_cards': deck_cards,
        'total': total,
        'is_owner': is_owner,
        'is_saved': is_saved,
    })


@login_required
def deck_edit(request, deck_id):
    deck = get_object_or_404(Deck, id=deck_id, user=request.user)
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if name:
            deck.name = name
            deck.save()
        return redirect('deck_edit', deck_id=deck.id)

    deck_cards = (
        deck.deck_cards
        .select_related('owned_card__card__card_set')
        .annotate(
            type_order=_card_type_order('owned_card__card__'),
            side_order=_side_order('owned_card__card__'),
        )
        .order_by('side_order', 'type_order', 'owned_card__card__name')
    )
    in_deck_ids = set(deck_cards.values_list('owned_card_id', flat=True))
    owned = (
        OwnedCard.objects
        .filter(user=request.user)
        .filter(Q(copies_bb__gt=0) | Q(copies_wb__gt=0))
        .select_related('card__card_set')
        .annotate(
            type_order=_card_type_order('card__'),
            side_order=_side_order('card__'),
        )
        .order_by('side_order', 'type_order', 'card__name')
    )
    total = deck_cards.aggregate(t=Sum('quantity'))['t'] or 0
    return render(request, 'swccgdb/decks/edit.html', {
        'deck': deck,
        'deck_cards': deck_cards,
        'owned': owned,
        'in_deck_ids': in_deck_ids,
        'total': total,
    })


@login_required
def deck_delete(request, deck_id):
    deck = get_object_or_404(Deck, id=deck_id, user=request.user)
    if request.method == 'POST':
        deck.delete()
    return redirect('deck_list')


@login_required
def deck_toggle_public(request, deck_id):
    deck = get_object_or_404(Deck, id=deck_id, user=request.user)
    if request.method == 'POST':
        deck.is_public = not deck.is_public
        deck.save()
    return redirect('deck_edit', deck_id=deck_id)


@login_required
def deck_save_toggle(request, deck_id):
    deck = get_object_or_404(Deck, id=deck_id)
    if deck.user == request.user:
        messages.error(request, "You can't save your own deck.")
        return redirect('deck_detail', deck_id=deck_id)
    if not deck.is_public:
        raise PermissionDenied
    if request.method == 'POST':
        if deck.saved_by.filter(id=request.user.id).exists():
            deck.saved_by.remove(request.user)
        else:
            deck.saved_by.add(request.user)
    return redirect(request.POST.get('next', 'deck_list'))


@login_required
def deck_add_card(request, deck_id):
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    deck = get_object_or_404(Deck, id=deck_id, user=request.user)
    owned_card = get_object_or_404(OwnedCard, id=request.POST.get('owned_card_id'), user=request.user)

    if DeckCard.objects.filter(deck=deck, owned_card=owned_card).exists():
        return JsonResponse({'error': 'Card already in deck'}, status=400)
    if owned_card.copies < 1:
        return JsonResponse({'error': f'You own no copies of {owned_card.card.name}'}, status=400)

    current_total = deck.deck_cards.aggregate(t=Sum('quantity'))['t'] or 0
    if current_total >= 60:
        return JsonResponse({'error': 'Deck is already at 60 cards'}, status=400)

    DeckCard.objects.create(deck=deck, owned_card=owned_card, quantity=1)
    return JsonResponse({'success': True, 'total': current_total + 1})


@login_required
def deck_remove_card(request, deck_id):
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    deck = get_object_or_404(Deck, id=deck_id, user=request.user)
    DeckCard.objects.filter(deck=deck, owned_card_id=request.POST.get('owned_card_id')).delete()
    total = deck.deck_cards.aggregate(t=Sum('quantity'))['t'] or 0
    return JsonResponse({'success': True, 'total': total})


@login_required
def deck_update_card(request, deck_id):
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    deck = get_object_or_404(Deck, id=deck_id, user=request.user)
    try:
        quantity = int(request.POST.get('quantity', 1))
    except (ValueError, TypeError):
        return JsonResponse({'error': 'Invalid quantity'}, status=400)
    if quantity < 1:
        return JsonResponse({'error': 'Quantity must be at least 1'}, status=400)

    deck_card = get_object_or_404(DeckCard, deck=deck, owned_card_id=request.POST.get('owned_card_id'))
    other_total = deck.deck_cards.exclude(id=deck_card.id).aggregate(t=Sum('quantity'))['t'] or 0
    if other_total + quantity > 60:
        return JsonResponse({'error': f'Deck cannot exceed 60 cards'}, status=400)
    if quantity > deck_card.owned_card.copies:
        return JsonResponse({'error': f'You only own {deck_card.owned_card.copies} copy/copies'}, status=400)

    deck_card.quantity = quantity
    deck_card.save()
    return JsonResponse({'success': True, 'total': other_total + quantity, 'quantity': quantity})

